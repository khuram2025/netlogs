"""
Project and Communication Matrix view routes.
"""

from datetime import datetime
from typing import Optional, List
from fastapi import APIRouter, Depends, Request, Form, Query, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import select, func, desc
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..core.auth import get_current_user
from ..db.database import get_db
from ..models.project import Project, CommunicationMatrixEntry, ProjectStatus, ConnectionType
from ..services.policy_builder_service import PolicyBuilderService, PolicyData
from ..schemas.project import (
    ProjectCreate, ProjectUpdate, ProjectResponse, ProjectDetailResponse,
    CommunicationMatrixEntryCreate, CommunicationMatrixEntryUpdate, CommunicationMatrixEntryResponse
)

router = APIRouter(tags=["projects"])

templates = Jinja2Templates(directory="fastapi_app/templates")


# ============================================================
# Project List & CRUD
# ============================================================

@router.get("/projects/", response_class=HTMLResponse, name="project_list")
async def project_list(
    request: Request,
    status: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """List all projects with filtering options."""
    query = select(Project).options(selectinload(Project.communication_entries))

    # Apply filters
    if status:
        query = query.where(Project.status == status.upper())
    if search:
        query = query.where(
            (Project.name.ilike(f"%{search}%")) |
            (Project.owner.ilike(f"%{search}%")) |
            (Project.description.ilike(f"%{search}%"))
        )

    query = query.order_by(desc(Project.updated_at))

    result = await db.execute(query)
    projects = result.scalars().all()

    # Get summary statistics
    stats_query = select(
        func.count(Project.id).label('total'),
        func.count(Project.id).filter(Project.status == ProjectStatus.ACTIVE).label('active'),
        func.count(Project.id).filter(Project.status == ProjectStatus.INACTIVE).label('inactive'),
    )
    stats_result = await db.execute(stats_query)
    stats_row = stats_result.one()

    # Get total communication entries
    entry_count_query = select(func.count(CommunicationMatrixEntry.id))
    entry_count_result = await db.execute(entry_count_query)
    total_entries = entry_count_result.scalar() or 0

    stats = {
        'total': stats_row.total,
        'active': stats_row.active,
        'inactive': stats_row.inactive,
        'total_entries': total_entries,
    }

    return templates.TemplateResponse("projects/project_list.html", {
        "request": request,
        "current_user": getattr(request.state, "current_user", None),
        "unread_alert_count": 0,
        "projects": projects,
        "stats": stats,
        "status_filter": status,
        "search": search,
        "status_choices": ProjectStatus.CHOICES,
    })


@router.get("/projects/new/", response_class=HTMLResponse, name="project_new")
async def project_new(request: Request):
    """Show create project form."""
    return templates.TemplateResponse("projects/project_form.html", {
        "request": request,
        "current_user": getattr(request.state, "current_user", None),
        "unread_alert_count": 0,
        "project": None,
        "status_choices": ProjectStatus.CHOICES,
        "is_edit": False,
    })


@router.post("/projects/new/", name="project_create")
async def project_create(
    name: str = Form(...),
    owner: str = Form(...),
    resources: Optional[str] = Form(None),
    location: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    status: str = Form("OTHER"),
    db: AsyncSession = Depends(get_db),
):
    """Create a new project."""
    project = Project(
        name=name.strip(),
        owner=owner.strip(),
        resources=resources.strip() if resources else None,
        location=location.strip() if location else None,
        description=description.strip() if description else None,
        status=status.upper(),
    )
    db.add(project)
    await db.commit()
    await db.refresh(project)

    return RedirectResponse(url=f"/projects/{project.id}/", status_code=303)


@router.get("/projects/{project_id}/", response_class=HTMLResponse, name="project_detail")
async def project_detail(
    request: Request,
    project_id: int,
    connection_type: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """Project detail page with communication matrix."""
    result = await db.execute(
        select(Project)
        .options(selectinload(Project.communication_entries))
        .where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()

    if not project:
        return RedirectResponse(url="/projects/", status_code=303)

    # Filter entries
    entries = project.communication_entries
    if connection_type:
        entries = [e for e in entries if e.connection_type == connection_type.upper()]
    if search:
        search_lower = search.lower()
        entries = [e for e in entries if (
            search_lower in e.source_ip.lower() or
            search_lower in e.destination_ip.lower() or
            search_lower in e.destination_port.lower() or
            (e.source_hostname and search_lower in e.source_hostname.lower()) or
            (e.destination_hostname and search_lower in e.destination_hostname.lower()) or
            (e.description and search_lower in e.description.lower())
        )]

    # Sort entries by creation date (newest first)
    entries = sorted(entries, key=lambda x: x.created_at, reverse=True)

    # Get entry statistics
    all_entries = project.communication_entries
    entry_stats = {
        'total': len(all_entries),
        'permanent': len([e for e in all_entries if e.connection_type == ConnectionType.PERMANENT]),
        'temporary': len([e for e in all_entries if e.connection_type == ConnectionType.TEMPORARY]),
        'active': len([e for e in all_entries if e.is_active]),
    }

    return templates.TemplateResponse("projects/project_detail.html", {
        "request": request,
        "current_user": getattr(request.state, "current_user", None),
        "unread_alert_count": 0,
        "project": project,
        "entries": entries,
        "entry_stats": entry_stats,
        "connection_type_filter": connection_type,
        "search": search,
        "connection_type_choices": ConnectionType.CHOICES,
        "protocol_choices": [('TCP', 'TCP'), ('UDP', 'UDP'), ('ICMP', 'ICMP'), ('ANY', 'Any'), ('TCP/UDP', 'TCP/UDP')],
    })


@router.get("/projects/{project_id}/edit/", response_class=HTMLResponse, name="project_edit")
async def project_edit(
    request: Request,
    project_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Show edit project form."""
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()

    if not project:
        return RedirectResponse(url="/projects/", status_code=303)

    return templates.TemplateResponse("projects/project_form.html", {
        "request": request,
        "current_user": getattr(request.state, "current_user", None),
        "unread_alert_count": 0,
        "project": project,
        "status_choices": ProjectStatus.CHOICES,
        "is_edit": True,
    })


@router.post("/projects/{project_id}/edit/", name="project_update")
async def project_update(
    project_id: int,
    name: str = Form(...),
    owner: str = Form(...),
    resources: Optional[str] = Form(None),
    location: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    status: str = Form("OTHER"),
    db: AsyncSession = Depends(get_db),
):
    """Update a project."""
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()

    if not project:
        return RedirectResponse(url="/projects/", status_code=303)

    project.name = name.strip()
    project.owner = owner.strip()
    project.resources = resources.strip() if resources else None
    project.location = location.strip() if location else None
    project.description = description.strip() if description else None
    project.status = status.upper()

    await db.commit()

    return RedirectResponse(url=f"/projects/{project_id}/", status_code=303)


@router.post("/projects/{project_id}/delete/", name="project_delete")
async def project_delete(
    project_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Delete a project and all its communication entries."""
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()

    if project:
        await db.delete(project)
        await db.commit()

    return RedirectResponse(url="/projects/", status_code=303)


# ============================================================
# Communication Matrix Entry CRUD
# ============================================================

@router.post("/projects/{project_id}/entries/add/", name="entry_add")
async def entry_add(
    project_id: int,
    source_ip: str = Form(...),
    source_hostname: Optional[str] = Form(None),
    destination_ip: str = Form(...),
    destination_hostname: Optional[str] = Form(None),
    destination_port: str = Form(...),
    protocol: str = Form("TCP"),
    connection_type: str = Form("PERMANENT"),
    description: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """Add a new communication matrix entry."""
    # Verify project exists
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()

    if not project:
        return RedirectResponse(url="/projects/", status_code=303)

    entry = CommunicationMatrixEntry(
        project_id=project_id,
        source_ip=source_ip.strip(),
        source_hostname=source_hostname.strip() if source_hostname else None,
        destination_ip=destination_ip.strip(),
        destination_hostname=destination_hostname.strip() if destination_hostname else None,
        destination_port=destination_port.strip(),
        protocol=protocol.upper(),
        connection_type=connection_type.upper(),
        description=description.strip() if description else None,
        is_active=True,
    )
    db.add(entry)
    await db.commit()

    return RedirectResponse(url=f"/projects/{project_id}/", status_code=303)


@router.post("/projects/{project_id}/entries/{entry_id}/update/", name="entry_update")
async def entry_update(
    project_id: int,
    entry_id: int,
    source_ip: str = Form(...),
    source_hostname: Optional[str] = Form(None),
    destination_ip: str = Form(...),
    destination_hostname: Optional[str] = Form(None),
    destination_port: str = Form(...),
    protocol: str = Form("TCP"),
    connection_type: str = Form("PERMANENT"),
    description: Optional[str] = Form(None),
    is_active: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    """Update a communication matrix entry."""
    result = await db.execute(
        select(CommunicationMatrixEntry)
        .where(CommunicationMatrixEntry.id == entry_id, CommunicationMatrixEntry.project_id == project_id)
    )
    entry = result.scalar_one_or_none()

    if not entry:
        return RedirectResponse(url=f"/projects/{project_id}/", status_code=303)

    entry.source_ip = source_ip.strip()
    entry.source_hostname = source_hostname.strip() if source_hostname else None
    entry.destination_ip = destination_ip.strip()
    entry.destination_hostname = destination_hostname.strip() if destination_hostname else None
    entry.destination_port = destination_port.strip()
    entry.protocol = protocol.upper()
    entry.connection_type = connection_type.upper()
    entry.description = description.strip() if description else None
    entry.is_active = is_active == "on"

    await db.commit()

    return RedirectResponse(url=f"/projects/{project_id}/", status_code=303)


@router.post("/projects/{project_id}/entries/{entry_id}/delete/", name="entry_delete")
async def entry_delete(
    project_id: int,
    entry_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Delete a communication matrix entry."""
    result = await db.execute(
        select(CommunicationMatrixEntry)
        .where(CommunicationMatrixEntry.id == entry_id, CommunicationMatrixEntry.project_id == project_id)
    )
    entry = result.scalar_one_or_none()

    if entry:
        await db.delete(entry)
        await db.commit()

    return RedirectResponse(url=f"/projects/{project_id}/", status_code=303)


@router.post("/projects/{project_id}/entries/{entry_id}/toggle/", name="entry_toggle")
async def entry_toggle(
    project_id: int,
    entry_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Toggle active status of a communication matrix entry."""
    result = await db.execute(
        select(CommunicationMatrixEntry)
        .where(CommunicationMatrixEntry.id == entry_id, CommunicationMatrixEntry.project_id == project_id)
    )
    entry = result.scalar_one_or_none()

    if entry:
        entry.is_active = not entry.is_active
        await db.commit()

    return RedirectResponse(url=f"/projects/{project_id}/", status_code=303)


# ============================================================
# API Endpoints (JSON)
# ============================================================

from fastapi import File, UploadFile
from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


@router.get("/api/projects/template/", name="api_project_template")
async def api_project_template():
    """Download a demo Excel template for importing communication matrix entries."""
    # Create workbook
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    centered = Alignment(horizontal='center', vertical='center')

    # ===== Instructions Sheet =====
    ws_help = wb.active
    ws_help.title = "Instructions"

    instructions = [
        ["Communication Matrix Import Template", ""],
        ["", ""],
        ["How to use this template:", ""],
        ["1. Go to the 'Communication Matrix' sheet", ""],
        ["2. Fill in your communication rules starting from row 2", ""],
        ["3. Required columns: Source IP, Destination IP, Port", ""],
        ["4. Save the file and import it into your project", ""],
        ["", ""],
        ["Column Descriptions:", ""],
        ["Source IP", "IP address or CIDR of the source (e.g., 192.168.1.0/24, 10.0.0.1)"],
        ["Source Hostname", "Optional: Server name or description (e.g., web-server-01)"],
        ["Destination IP", "IP address or CIDR of the destination"],
        ["Destination Hostname", "Optional: Server name or description (e.g., db-master)"],
        ["Port", "Destination port(s) - single (443), range (8000-9000), or multiple (80,443,8080)"],
        ["Protocol", "TCP, UDP, ICMP, ANY, or TCP/UDP (default: TCP)"],
        ["Connection Type", "PERMANENT or TEMPORARY (default: PERMANENT)"],
        ["Description", "Optional: Description of this communication rule"],
        ["Active", "Yes or No (default: Yes)"],
        ["", ""],
        ["Supported IP Formats:", ""],
        ["Single IP", "192.168.1.100"],
        ["CIDR Notation", "192.168.1.0/24"],
        ["IP Range", "192.168.1.1-192.168.1.100"],
        ["Multiple IPs", "192.168.1.1, 192.168.1.2, 192.168.1.3"],
        ["Any", "any"],
    ]

    for row_idx, (col1, col2) in enumerate(instructions, 1):
        ws_help.cell(row=row_idx, column=1, value=col1)
        ws_help.cell(row=row_idx, column=2, value=col2)
        if row_idx == 1:
            ws_help.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
        elif col1 and ":" in col1:
            ws_help.cell(row=row_idx, column=1).font = Font(bold=True)

    ws_help.column_dimensions['A'].width = 25
    ws_help.column_dimensions['B'].width = 60

    # ===== Communication Matrix Sheet =====
    ws_matrix = wb.create_sheet("Communication Matrix")

    # Headers
    headers = [
        "Source IP", "Source Hostname", "Destination IP", "Destination Hostname",
        "Port", "Protocol", "Connection Type", "Description", "Active"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws_matrix.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = centered

    # Demo data
    demo_data = [
        ["192.168.10.0/24", "web-servers", "10.0.50.10", "db-master-01", "5432", "TCP", "PERMANENT", "Web servers to PostgreSQL database", "Yes"],
        ["172.16.1.0/24", "app-servers", "10.0.50.20", "redis-cluster", "6379", "TCP", "PERMANENT", "Application servers to Redis cache", "Yes"],
        ["10.0.100.5", "admin-workstation", "192.168.10.0/24", "web-servers", "22,443", "TCP", "PERMANENT", "Admin SSH and HTTPS access to web servers", "Yes"],
        ["192.168.1.100", "monitoring-server", "10.0.0.0/8", "all-servers", "161", "UDP", "PERMANENT", "SNMP monitoring", "Yes"],
        ["any", "", "203.0.113.50", "public-web", "80,443", "TCP", "PERMANENT", "Public access to web server", "Yes"],
        ["10.10.10.0/24", "dev-network", "10.20.20.100", "test-db", "3306", "TCP", "TEMPORARY", "Development access to test database", "Yes"],
        ["172.16.0.1-172.16.0.10", "load-balancers", "192.168.100.0/24", "backend-pool", "8080-8089", "TCP", "PERMANENT", "Load balancer to backend services", "Yes"],
        ["192.168.50.0/24", "office-network", "10.0.0.1", "dns-server", "53", "UDP", "PERMANENT", "DNS queries", "Yes"],
    ]

    for row_idx, row_data in enumerate(demo_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_matrix.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.fill = example_fill

    # Adjust column widths
    column_widths = [22, 20, 22, 20, 18, 12, 18, 45, 10]
    for col_idx, width in enumerate(column_widths, 1):
        ws_matrix.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws_matrix.freeze_panes = "A2"

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=communication_matrix_template.xlsx"}
    )


@router.get("/api/projects/", response_model=List[ProjectResponse], name="api_project_list")
async def api_project_list(
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """Get all projects as JSON."""
    query = select(Project).options(selectinload(Project.communication_entries))

    if status:
        query = query.where(Project.status == status.upper())

    query = query.order_by(desc(Project.updated_at))

    result = await db.execute(query)
    projects = result.scalars().all()

    return projects


@router.get("/api/projects/{project_id}/", response_model=ProjectDetailResponse, name="api_project_detail")
async def api_project_detail(
    project_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Get project details with communication entries as JSON."""
    result = await db.execute(
        select(Project)
        .options(selectinload(Project.communication_entries))
        .where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    return project


@router.get("/api/projects/{project_id}/entries/", response_model=List[CommunicationMatrixEntryResponse], name="api_entry_list")
async def api_entry_list(
    project_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Get all communication matrix entries for a project as JSON."""
    result = await db.execute(
        select(CommunicationMatrixEntry)
        .where(CommunicationMatrixEntry.project_id == project_id)
        .order_by(desc(CommunicationMatrixEntry.created_at))
    )
    entries = result.scalars().all()

    return entries


# ============================================================
# Policy Builder Endpoints
# ============================================================

def parse_ip_input(ip_input: str) -> List[str]:
    """
    Parse IP input supporting multiple formats:
    - Single IP: 192.168.1.1
    - CIDR: 192.168.1.0/24
    - Range: 192.168.1.1-192.168.1.100
    - Multiple IPs: 192.168.1.1, 10.0.0.1, 172.16.0.1
    - Mixed: 192.168.1.0/24, 10.0.0.1-10.0.0.50

    Returns list of individual IP expressions for policy building.
    """
    import ipaddress

    ip_input = ip_input.strip()
    if ip_input.lower() == 'any':
        return ['any']

    results = []

    # Split by comma for multiple entries
    parts = [p.strip() for p in ip_input.split(',')]

    for part in parts:
        part = part.strip()
        if not part:
            continue

        # Check if it's a range (contains '-' but not in CIDR)
        if '-' in part and '/' not in part:
            # IP range format: 192.168.1.1-192.168.1.100 or 192.168.1.1-100
            try:
                if part.count('.') > 3:
                    # Full range: 192.168.1.1-192.168.1.100
                    start_ip, end_ip = part.split('-')
                    start = ipaddress.ip_address(start_ip.strip())
                    end = ipaddress.ip_address(end_ip.strip())
                else:
                    # Short range: 192.168.1.1-100
                    base_part, end_octet = part.rsplit('-', 1)
                    start = ipaddress.ip_address(base_part.strip())
                    # Build end IP by replacing last octet
                    start_parts = str(start).split('.')
                    start_parts[-1] = end_octet.strip()
                    end = ipaddress.ip_address('.'.join(start_parts))

                # For ranges, we'll return the range notation for FortiGate
                results.append(f"{start}-{end}")
            except ValueError:
                # If parsing fails, add as-is
                results.append(part)
        else:
            # Single IP or CIDR
            results.append(part)

    return results if results else [ip_input]


def generate_policy_cli_for_entry(
    entry: CommunicationMatrixEntry,
    project_name: str,
    policy_index: int = 0,
    vendor: str = "fortinet"
) -> dict:
    """Generate firewall CLI for a single communication matrix entry.

    Args:
        entry: Communication matrix entry
        project_name: Name of the project
        policy_index: Index for naming policies
        vendor: 'fortinet' or 'paloalto'
    """
    vendor = vendor.lower() if vendor else "fortinet"

    if vendor == "paloalto":
        return generate_paloalto_policy_cli_for_entry(entry, project_name, policy_index)
    else:
        return generate_fortinet_policy_cli_for_entry(entry, project_name, policy_index)


def sanitize_name_for_firewall(name: str, max_length: int = 63) -> str:
    """
    Sanitize a name for use as a firewall object name.
    - Replace spaces and special characters with underscores
    - Ensure it starts with a letter or underscore
    - Limit to max_length characters
    """
    import re
    # Replace spaces and special chars with underscores
    safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', name)
    # Remove consecutive underscores
    safe_name = re.sub(r'_+', '_', safe_name)
    # Remove leading/trailing underscores
    safe_name = safe_name.strip('_')
    # Ensure it doesn't start with a number
    if safe_name and safe_name[0].isdigit():
        safe_name = '_' + safe_name
    return safe_name[:max_length] if safe_name else 'addr'


def generate_fortinet_policy_cli_for_entry(
    entry: CommunicationMatrixEntry,
    project_name: str,
    policy_index: int = 0
) -> dict:
    """Generate FortiGate CLI for a single communication matrix entry."""

    cli_sections = []
    all_cli = []

    # Parse source and destination IPs
    source_ips = parse_ip_input(entry.source_ip)
    dest_ips = parse_ip_input(entry.destination_ip)

    # Parse ports (support ranges and multiple ports)
    ports = []
    port_input = entry.destination_port.strip()
    if port_input.lower() == 'any':
        ports = ['any']
    else:
        for port_part in port_input.split(','):
            port_part = port_part.strip()
            if port_part:
                ports.append(port_part)

    # Map protocol
    protocol = entry.protocol.upper()
    proto_map = {'TCP': 'tcp', 'UDP': 'udp', 'ICMP': 'icmp', 'ANY': 'any', 'TCP/UDP': 'tcp'}
    proto = proto_map.get(protocol, 'tcp')

    # Determine address object names based on hostname or IP
    # Use hostname if available, otherwise use IP-based naming
    src_hostname = entry.source_hostname.strip() if entry.source_hostname else None
    dst_hostname = entry.destination_hostname.strip() if entry.destination_hostname else None

    # Generate address objects for source IPs
    src_addr_names = []
    for i, src_ip in enumerate(source_ips):
        if src_ip.lower() == 'any':
            src_addr_names.append('all')
        else:
            # Use hostname if available, otherwise use project-based naming
            if src_hostname and len(source_ips) == 1:
                # Single IP with hostname - use hostname as name
                addr_name = sanitize_name_for_firewall(src_hostname, 63)
            elif src_hostname and len(source_ips) > 1:
                # Multiple IPs with hostname - append index
                addr_name = sanitize_name_for_firewall(f"{src_hostname}_{i}", 63)
            else:
                # No hostname - use IP-based naming
                safe_ip = src_ip.replace('.', '_').replace('/', '_').replace('-', '_')
                addr_name = sanitize_name_for_firewall(f"src_{safe_ip}", 63)

            # Check if it's a range
            if '-' in src_ip and '/' not in src_ip:
                comment = src_hostname or entry.description or 'communication matrix'
                cli_sections.append(f'''config firewall address
    edit "{addr_name}"
        set type iprange
        set start-ip {src_ip.split('-')[0]}
        set end-ip {src_ip.split('-')[1]}
        set comment "{comment}"
    next
end''')
            else:
                cli_sections.append(PolicyBuilderService.build_address_object_cli(src_ip, addr_name))
            src_addr_names.append(addr_name)

    # Generate address objects for destination IPs
    dst_addr_names = []
    for i, dst_ip in enumerate(dest_ips):
        if dst_ip.lower() == 'any':
            dst_addr_names.append('all')
        else:
            # Use hostname if available, otherwise use IP-based naming
            if dst_hostname and len(dest_ips) == 1:
                # Single IP with hostname - use hostname as name
                addr_name = sanitize_name_for_firewall(dst_hostname, 63)
            elif dst_hostname and len(dest_ips) > 1:
                # Multiple IPs with hostname - append index
                addr_name = sanitize_name_for_firewall(f"{dst_hostname}_{i}", 63)
            else:
                # No hostname - use IP-based naming
                safe_ip = dst_ip.replace('.', '_').replace('/', '_').replace('-', '_')
                addr_name = sanitize_name_for_firewall(f"dst_{safe_ip}", 63)

            if '-' in dst_ip and '/' not in dst_ip:
                comment = dst_hostname or entry.description or 'communication matrix'
                cli_sections.append(f'''config firewall address
    edit "{addr_name}"
        set type iprange
        set start-ip {dst_ip.split('-')[0]}
        set end-ip {dst_ip.split('-')[1]}
        set comment "{comment}"
    next
end''')
            else:
                cli_sections.append(PolicyBuilderService.build_address_object_cli(dst_ip, addr_name))
            dst_addr_names.append(addr_name)

    # Generate service objects for non-standard ports
    service_names = []
    for port in ports:
        if port.lower() == 'any':
            service_names.append('ALL')
        elif '-' in port:
            # Port range
            svc_name = f"svc_{proto.upper()}_{port.replace('-', '_')}"
            cli_sections.append(f'''config firewall service custom
    edit "{svc_name}"
        set {proto}-portrange {port}
    next
end''')
            service_names.append(svc_name)
        else:
            try:
                port_int = int(port)
                svc_name = PolicyBuilderService.get_service_name(port_int, proto)
                svc_cli = PolicyBuilderService.build_service_object_cli(port_int, proto)
                if svc_cli:
                    cli_sections.append(svc_cli)
                service_names.append(svc_name)
            except ValueError:
                service_names.append(port)

    # Generate policy
    policy_name = f"{project_name}_{policy_index}"[:35]
    src_addr_str = '" "'.join(src_addr_names)
    dst_addr_str = '" "'.join(dst_addr_names)
    service_str = '" "'.join(service_names) if service_names else 'ALL'

    comment = entry.description or f"Communication matrix rule: {entry.source_ip} -> {entry.destination_ip}:{entry.destination_port}"
    comment = comment[:255]  # FortiGate comment limit

    policy_cli = f'''config firewall policy
    edit 0
        set name "{policy_name}"
        set srcintf "any"
        set dstintf "any"
        set srcaddr "{src_addr_str}"
        set dstaddr "{dst_addr_str}"
        set action accept
        set schedule "always"
        set service "{service_str}"
        set logtraffic all
        set comments "{comment}"
    next
end'''
    cli_sections.append(policy_cli)

    return {
        'entry_id': entry.id,
        'source_ip': entry.source_ip,
        'destination_ip': entry.destination_ip,
        'destination_port': entry.destination_port,
        'protocol': entry.protocol,
        'connection_type': entry.connection_type,
        'description': entry.description,
        'cli': '\n\n'.join(cli_sections),
        'vendor': 'fortinet'
    }


def generate_paloalto_policy_cli_for_entry(
    entry: CommunicationMatrixEntry,
    project_name: str,
    policy_index: int = 0
) -> dict:
    """Generate Palo Alto Networks CLI for a single communication matrix entry."""

    cli_sections = []

    # Parse source and destination IPs
    source_ips = parse_ip_input(entry.source_ip)
    dest_ips = parse_ip_input(entry.destination_ip)

    # Parse ports (support ranges and multiple ports)
    ports = []
    port_input = entry.destination_port.strip()
    if port_input.lower() == 'any':
        ports = ['any']
    else:
        for port_part in port_input.split(','):
            port_part = port_part.strip()
            if port_part:
                ports.append(port_part)

    # Map protocol
    protocol = entry.protocol.upper()
    proto_map = {'TCP': 'tcp', 'UDP': 'udp', 'ICMP': 'icmp', 'ANY': 'any', 'TCP/UDP': 'tcp'}
    proto = proto_map.get(protocol, 'tcp')

    # Determine address object names based on hostname or IP
    src_hostname = entry.source_hostname.strip() if entry.source_hostname else None
    dst_hostname = entry.destination_hostname.strip() if entry.destination_hostname else None

    # Helper to sanitize for Palo Alto (uses dashes instead of underscores)
    def sanitize_paloalto_name(name: str, max_len: int = 63) -> str:
        import re
        safe = re.sub(r'[^a-zA-Z0-9_-]', '-', name)
        safe = re.sub(r'-+', '-', safe)
        safe = safe.strip('-')
        if safe and safe[0].isdigit():
            safe = 'addr-' + safe
        return safe[:max_len] if safe else 'addr'

    # Generate address objects for source IPs
    src_addr_names = []

    for i, src_ip in enumerate(source_ips):
        if src_ip.lower() == 'any':
            src_addr_names.append('any')
        else:
            # Use hostname if available, otherwise use IP-based naming
            if src_hostname and len(source_ips) == 1:
                addr_name = sanitize_paloalto_name(src_hostname, 63)
            elif src_hostname and len(source_ips) > 1:
                addr_name = sanitize_paloalto_name(f"{src_hostname}-{i}", 63)
            else:
                safe_ip = src_ip.replace('.', '-').replace('/', '-')
                addr_name = sanitize_paloalto_name(f"src-{safe_ip}", 63)

            # Check if it's a range
            if '-' in src_ip and '/' not in src_ip:
                start_ip, end_ip = src_ip.split('-')
                cli_sections.append(f'set address {addr_name} ip-range {start_ip}-{end_ip}')
            elif '/' in src_ip:
                cli_sections.append(f'set address {addr_name} ip-netmask {src_ip}')
            else:
                cli_sections.append(f'set address {addr_name} ip-netmask {src_ip}/32')
            src_addr_names.append(addr_name)

    # Generate address objects for destination IPs
    dst_addr_names = []
    for i, dst_ip in enumerate(dest_ips):
        if dst_ip.lower() == 'any':
            dst_addr_names.append('any')
        else:
            # Use hostname if available, otherwise use IP-based naming
            if dst_hostname and len(dest_ips) == 1:
                addr_name = sanitize_paloalto_name(dst_hostname, 63)
            elif dst_hostname and len(dest_ips) > 1:
                addr_name = sanitize_paloalto_name(f"{dst_hostname}-{i}", 63)
            else:
                safe_ip = dst_ip.replace('.', '-').replace('/', '-')
                addr_name = sanitize_paloalto_name(f"dst-{safe_ip}", 63)

            if '-' in dst_ip and '/' not in dst_ip:
                start_ip, end_ip = dst_ip.split('-')
                cli_sections.append(f'set address {addr_name} ip-range {start_ip}-{end_ip}')
            elif '/' in dst_ip:
                cli_sections.append(f'set address {addr_name} ip-netmask {dst_ip}')
            else:
                cli_sections.append(f'set address {addr_name} ip-netmask {dst_ip}/32')
            dst_addr_names.append(addr_name)

    # Palo Alto service mappings for well-known ports
    paloalto_services = {
        (22, 'tcp'): 'ssh', (80, 'tcp'): 'web-browsing', (443, 'tcp'): 'ssl',
        (53, 'tcp'): 'dns', (53, 'udp'): 'dns', (21, 'tcp'): 'ftp',
        (25, 'tcp'): 'smtp', (110, 'tcp'): 'pop3', (143, 'tcp'): 'imap',
        (3389, 'tcp'): 'ms-rdp', (3306, 'tcp'): 'mysql', (5432, 'tcp'): 'postgres',
        (123, 'udp'): 'ntp', (161, 'udp'): 'snmp', (514, 'udp'): 'syslog',
    }

    # Generate service objects for non-standard ports
    service_names = []
    application = 'any'

    for port in ports:
        if port.lower() == 'any':
            service_names.append('any')
        elif '-' in port:
            # Port range
            svc_name = f"svc-{proto}-{port.replace('-', '-to-')}"
            cli_sections.append(f'set service {svc_name} protocol {proto} port {port}')
            service_names.append(svc_name)
        else:
            try:
                port_int = int(port)
                # Check for standard Palo Alto service
                if (port_int, proto) in paloalto_services:
                    application = paloalto_services[(port_int, proto)]
                    service_names.append('application-default')
                else:
                    svc_name = f"svc-{proto}-{port}"
                    cli_sections.append(f'set service {svc_name} protocol {proto} port {port}')
                    service_names.append(svc_name)
            except ValueError:
                service_names.append(port)

    # Generate security policy rule
    safe_project = sanitize_paloalto_name(project_name, 20)
    policy_name = f"{safe_project}-rule-{policy_index}"[:31]
    src_addr_str = ' '.join(src_addr_names)
    dst_addr_str = ' '.join(dst_addr_names)

    # Use first service or 'any'
    if service_names:
        service_str = service_names[0]
    else:
        service_str = 'any'

    description = entry.description or f"Rule: {entry.source_ip} -> {entry.destination_ip}:{entry.destination_port}"
    description = description[:255].replace('"', "'")

    # Build the security rule
    rule_cli = f'''set rulebase security rules {policy_name} from any
set rulebase security rules {policy_name} to any
set rulebase security rules {policy_name} source {src_addr_str}
set rulebase security rules {policy_name} destination {dst_addr_str}
set rulebase security rules {policy_name} application {application}
set rulebase security rules {policy_name} service {service_str}
set rulebase security rules {policy_name} action allow
set rulebase security rules {policy_name} log-end yes
set rulebase security rules {policy_name} description "{description}"'''

    cli_sections.append(rule_cli)

    return {
        'entry_id': entry.id,
        'source_ip': entry.source_ip,
        'destination_ip': entry.destination_ip,
        'destination_port': entry.destination_port,
        'protocol': entry.protocol,
        'connection_type': entry.connection_type,
        'description': entry.description,
        'cli': '\n'.join(cli_sections),
        'vendor': 'paloalto'
    }


@router.get("/api/projects/{project_id}/entries/{entry_id}/policy/", name="api_entry_policy")
async def api_entry_policy(
    project_id: int,
    entry_id: int,
    vendor: str = "fortinet",
    db: AsyncSession = Depends(get_db),
):
    """Generate firewall CLI policy for a single communication matrix entry.

    Args:
        vendor: 'fortinet' (default) or 'paloalto'
    """
    # Get project
    project_result = await db.execute(select(Project).where(Project.id == project_id))
    project = project_result.scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Get entry
    entry_result = await db.execute(
        select(CommunicationMatrixEntry)
        .where(CommunicationMatrixEntry.id == entry_id, CommunicationMatrixEntry.project_id == project_id)
    )
    entry = entry_result.scalar_one_or_none()

    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    policy = generate_policy_cli_for_entry(entry, project.name, entry.id, vendor=vendor)

    return {
        "project_name": project.name,
        "project_id": project.id,
        "policy": policy
    }


@router.get("/api/projects/{project_id}/policy/", name="api_project_policy")
async def api_project_policy(
    project_id: int,
    active_only: bool = True,
    vendor: str = "fortinet",
    db: AsyncSession = Depends(get_db),
):
    """Generate firewall CLI policies for all communication matrix entries in a project.

    Args:
        vendor: 'fortinet' (default) or 'paloalto'
    """
    # Get project with entries
    result = await db.execute(
        select(Project)
        .options(selectinload(Project.communication_entries))
        .where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    entries = project.communication_entries
    if active_only:
        entries = [e for e in entries if e.is_active]

    policies = []
    all_cli_parts = []

    # Vendor-specific header
    vendor_name = "Palo Alto Networks" if vendor.lower() == "paloalto" else "FortiGate"

    # Add header comment
    all_cli_parts.append(f'''# ============================================
# {vendor_name} Policy Configuration
# Project: {project.name}
# Owner: {project.owner}
# Generated: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
# Total Rules: {len(entries)}
# ============================================
''')

    for idx, entry in enumerate(entries):
        policy = generate_policy_cli_for_entry(entry, project.name, idx + 1, vendor=vendor)
        policies.append(policy)
        all_cli_parts.append(f"# Rule {idx + 1}: {entry.source_ip} -> {entry.destination_ip}:{entry.destination_port}")
        all_cli_parts.append(policy['cli'])
        all_cli_parts.append('')

    # Add commit for Palo Alto
    if vendor.lower() == "paloalto":
        all_cli_parts.append('commit')

    return {
        "project_name": project.name,
        "project_id": project.id,
        "project_owner": project.owner,
        "total_entries": len(entries),
        "policies": policies,
        "combined_cli": '\n'.join(all_cli_parts),
        "vendor": vendor.lower()
    }


@router.get("/projects/{project_id}/policy/", response_class=HTMLResponse, name="project_policy_view")
async def project_policy_view(
    request: Request,
    project_id: int,
    vendor: str = "fortinet",
    db: AsyncSession = Depends(get_db),
):
    """View generated policies for a project.

    Args:
        vendor: 'fortinet' (default) or 'paloalto'
    """
    # Get project with entries
    result = await db.execute(
        select(Project)
        .options(selectinload(Project.communication_entries))
        .where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()

    if not project:
        return RedirectResponse(url="/projects/", status_code=303)

    entries = [e for e in project.communication_entries if e.is_active]

    policies = []
    all_cli_parts = []

    # Vendor-specific header
    vendor_name = "Palo Alto Networks" if vendor.lower() == "paloalto" else "FortiGate"

    # Add header
    import datetime
    all_cli_parts.append(f'''# ============================================
# {vendor_name} Policy Configuration
# Project: {project.name}
# Owner: {project.owner}
# Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
# Total Rules: {len(entries)}
# ============================================
''')

    for idx, entry in enumerate(entries):
        policy = generate_policy_cli_for_entry(entry, project.name, idx + 1, vendor=vendor)
        policies.append(policy)
        all_cli_parts.append(f"# Rule {idx + 1}: {entry.source_ip} -> {entry.destination_ip}:{entry.destination_port}")
        all_cli_parts.append(policy['cli'])
        all_cli_parts.append('')

    # Add commit for Palo Alto
    if vendor.lower() == "paloalto":
        all_cli_parts.append('commit')

    return templates.TemplateResponse("projects/project_policy.html", {
        "request": request,
        "current_user": getattr(request.state, "current_user", None),
        "unread_alert_count": 0,
        "project": project,
        "policies": policies,
        "combined_cli": '\n'.join(all_cli_parts),
        "total_entries": len(entries),
        "vendor": vendor.lower(),
        "vendor_name": vendor_name,
    })


# ============================================================
# Import/Export Endpoints
# ============================================================

@router.get("/api/projects/{project_id}/export/", name="api_project_export")
async def api_project_export(
    project_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Export project and communication matrix entries to Excel file."""
    # Get project with entries
    result = await db.execute(
        select(Project)
        .options(selectinload(Project.communication_entries))
        .where(Project.id == project_id)
    )
    project = result.scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Create workbook
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    centered = Alignment(horizontal='center', vertical='center')

    # ===== Project Info Sheet =====
    ws_info = wb.active
    ws_info.title = "Project Info"

    # Project details
    info_data = [
        ["Project Name", project.name],
        ["Owner", project.owner],
        ["Status", project.status],
        ["Location", project.location or ""],
        ["Resources", project.resources or ""],
        ["Description", project.description or ""],
        ["Created", project.created_at.strftime("%Y-%m-%d %H:%M:%S") if project.created_at else ""],
        ["Updated", project.updated_at.strftime("%Y-%m-%d %H:%M:%S") if project.updated_at else ""],
    ]

    for row_idx, (label, value) in enumerate(info_data, 1):
        ws_info.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_info.cell(row=row_idx, column=2, value=value)

    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 50

    # ===== Communication Matrix Sheet =====
    ws_matrix = wb.create_sheet("Communication Matrix")

    # Headers
    headers = [
        "Source IP", "Source Hostname", "Destination IP", "Destination Hostname",
        "Port", "Protocol", "Connection Type", "Description", "Active"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws_matrix.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = centered

    # Data rows
    entries = sorted(project.communication_entries, key=lambda x: x.created_at or x.id, reverse=True)
    for row_idx, entry in enumerate(entries, 2):
        row_data = [
            entry.source_ip,
            entry.source_hostname or "",
            entry.destination_ip,
            entry.destination_hostname or "",
            entry.destination_port,
            entry.protocol,
            entry.connection_type,
            entry.description or "",
            "Yes" if entry.is_active else "No"
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_matrix.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Adjust column widths
    column_widths = [20, 25, 20, 25, 15, 12, 15, 40, 10]
    for col_idx, width in enumerate(column_widths, 1):
        ws_matrix.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws_matrix.freeze_panes = "A2"

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename
    safe_name = project.name.replace(' ', '_').replace('/', '_')[:30]
    filename = f"{safe_name}_communication_matrix.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/projects/{project_id}/import/", name="project_import")
async def project_import(
    project_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Import communication matrix entries from Excel file."""
    # Verify project exists
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()

    if not project:
        return RedirectResponse(url="/projects/", status_code=303)

    # Check file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        return JSONResponse(
            {"success": False, "error": "Invalid file type. Please upload an Excel file (.xlsx)"},
            status_code=400
        )

    try:
        # Read file content
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))

        # Find the communication matrix sheet
        sheet_names = wb.sheetnames
        ws = None
        for name in ["Communication Matrix", "Matrix", "Rules", "Entries"]:
            if name in sheet_names:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active  # Use first sheet if no matching name

        # Parse headers (first row)
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val).lower().strip() if val else "")

        # Map headers to expected fields
        header_map = {
            'source ip': 'source_ip',
            'source_ip': 'source_ip',
            'src ip': 'source_ip',
            'srcip': 'source_ip',
            'source hostname': 'source_hostname',
            'source_hostname': 'source_hostname',
            'src hostname': 'source_hostname',
            'source host': 'source_hostname',
            'source server': 'source_hostname',
            'destination ip': 'destination_ip',
            'destination_ip': 'destination_ip',
            'dst ip': 'destination_ip',
            'dstip': 'destination_ip',
            'dest ip': 'destination_ip',
            'destination hostname': 'destination_hostname',
            'destination_hostname': 'destination_hostname',
            'dst hostname': 'destination_hostname',
            'destination host': 'destination_hostname',
            'destination server': 'destination_hostname',
            'port': 'destination_port',
            'destination port': 'destination_port',
            'destination_port': 'destination_port',
            'dst port': 'destination_port',
            'dstport': 'destination_port',
            'protocol': 'protocol',
            'proto': 'protocol',
            'connection type': 'connection_type',
            'connection_type': 'connection_type',
            'type': 'connection_type',
            'description': 'description',
            'desc': 'description',
            'comment': 'description',
            'comments': 'description',
            'active': 'is_active',
            'is_active': 'is_active',
            'enabled': 'is_active',
        }

        # Create column index mapping
        col_map = {}
        for idx, header in enumerate(headers):
            if header in header_map:
                col_map[header_map[header]] = idx + 1

        # Validate required columns
        required = ['source_ip', 'destination_ip', 'destination_port']
        missing = [r for r in required if r not in col_map]
        if missing:
            return JSONResponse(
                {"success": False, "error": f"Missing required columns: {', '.join(missing)}"},
                status_code=400
            )

        # Import rows
        imported_count = 0
        errors = []

        for row_idx in range(2, ws.max_row + 1):
            try:
                # Get cell values
                def get_val(field, default=""):
                    if field in col_map:
                        val = ws.cell(row=row_idx, column=col_map[field]).value
                        return str(val).strip() if val else default
                    return default

                source_ip = get_val('source_ip')
                destination_ip = get_val('destination_ip')
                destination_port = get_val('destination_port')

                # Skip empty rows
                if not source_ip or not destination_ip or not destination_port:
                    continue

                # Parse other fields
                source_hostname = get_val('source_hostname') or None
                destination_hostname = get_val('destination_hostname') or None
                protocol = get_val('protocol', 'TCP').upper()
                if protocol not in ['TCP', 'UDP', 'ICMP', 'ANY', 'TCP/UDP']:
                    protocol = 'TCP'

                connection_type = get_val('connection_type', 'PERMANENT').upper()
                if connection_type not in ['PERMANENT', 'TEMPORARY']:
                    connection_type = 'PERMANENT'

                description = get_val('description') or None

                is_active_str = get_val('is_active', 'yes').lower()
                is_active = is_active_str in ['yes', 'true', '1', 'y', 'active', 'enabled']

                # Create entry
                entry = CommunicationMatrixEntry(
                    project_id=project_id,
                    source_ip=source_ip,
                    source_hostname=source_hostname,
                    destination_ip=destination_ip,
                    destination_hostname=destination_hostname,
                    destination_port=destination_port,
                    protocol=protocol,
                    connection_type=connection_type,
                    description=description,
                    is_active=is_active,
                )
                db.add(entry)
                imported_count += 1

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        await db.commit()

        # Return to project page with success message
        return RedirectResponse(
            url=f"/projects/{project_id}/?imported={imported_count}",
            status_code=303
        )

    except Exception as e:
        return JSONResponse(
            {"success": False, "error": f"Error processing file: {str(e)}"},
            status_code=400
        )
